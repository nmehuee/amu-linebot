import os
import io
import re
import datetime
import traceback
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from flask import Flask, request, abort
from linebot import LineBotApi, WebhookHandler
from linebot.exceptions import InvalidSignatureError
from linebot.models import (
    MessageEvent, TextMessage, TextSendMessage,
    PostbackEvent, FlexSendMessage, BubbleContainer,
    BoxComponent, ButtonComponent, TextComponent,
    PostbackAction, URIAction, ImageComponent,
    SeparatorComponent
)
from supabase import create_client, Client

app = Flask(__name__)

CHANNEL_ACCESS_TOKEN = os.environ.get("CHANNEL_ACCESS_TOKEN")
CHANNEL_SECRET = os.environ.get("CHANNEL_SECRET")
OWNER_ID = os.environ.get("OWNER_ID")
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY")

line_bot_api = LineBotApi(CHANNEL_ACCESS_TOKEN)
handler = WebhookHandler(CHANNEL_SECRET)
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

user_states = {}

PRICE_PER_PACK = 200
FREE_SHIPPING_THRESHOLD = 2000
SHIPPING_FEE = 170
MIN_PACKS = 2
MAX_PACKS = 15


def calc_order(qty_a, qty_b):
    subtotal = (qty_a + qty_b) * PRICE_PER_PACK
    shipping = 0 if subtotal >= FREE_SHIPPING_THRESHOLD else SHIPPING_FEE
    total = subtotal + shipping
    return subtotal, shipping, total


def reset_state(user_id):
    user_states.pop(user_id, None)


def make_cancel_button():
    return ButtonComponent(
        action=PostbackAction(label="❌ 取消訂單", data="action=cancel"),
        style="secondary",
        color="#AAAAAA",
        margin="lg"
    )


def send_welcome(reply_token):
    bubble = BubbleContainer(
        direction="ltr",
        hero=ImageComponent(
            url="https://i.imgur.com/xxxxxx.jpg",
            size="full",
            aspect_ratio="20:13",
            aspect_mode="cover"
        ),
        body=BoxComponent(
            layout="vertical",
            contents=[
                TextComponent(text="🥟 A-MU 水餃", weight="bold", size="xl", color="#333333"),
                TextComponent(text="高麗菜黑豬 / 韭菜黑豬", size="sm", color="#888888", margin="sm"),
                SeparatorComponent(margin="md"),
                TextComponent(
                    text="每包 NT$200｜最少 2 包｜最多 15 包\n滿 $2000 免運，未滿加收 $170",
                    size="sm",
                    color="#555555",
                    margin="md",
                    wrap=True
                ),
            ]
        ),
        footer=BoxComponent(
            layout="vertical",
            contents=[
                ButtonComponent(
                    action=PostbackAction(label="🛒 開始訂購", data="action=start_order"),
                    style="primary",
                    color="#FF6B6B"
                )
            ]
        )
    )
    line_bot_api.reply_message(
        reply_token,
        FlexSendMessage(alt_text="歡迎來到 A-MU 水餃！", contents=bubble)
    )


def send_qty_a_flex(reply_token):
    buttons = []
    for i in range(0, 16):
        buttons.append(
            ButtonComponent(
                action=PostbackAction(label=str(i), data=f"action=qty_a&value={i}"),
                style="primary" if i > 0 else "secondary",
                color="#FF6B6B" if i > 0 else "#AAAAAA",
                height="sm"
            )
        )
    rows = []
    row = []
    for idx, btn in enumerate(buttons):
        row.append(btn)
        if len(row) == 4:
            rows.append(BoxComponent(layout="horizontal", contents=row, spacing="sm"))
            row = []
    if row:
        rows.append(BoxComponent(layout="horizontal", contents=row, spacing="sm"))

    bubble = BubbleContainer(
        body=BoxComponent(
            layout="vertical",
            contents=[
                TextComponent(text="高麗菜黑豬水餃", weight="bold", size="lg", color="#333333"),
                TextComponent(text="請選擇數量（包）", size="sm", color="#888888", margin="sm"),
                SeparatorComponent(margin="md"),
                *rows,
                make_cancel_button()
            ],
            spacing="md"
        )
    )
    line_bot_api.reply_message(
        reply_token,
        FlexSendMessage(alt_text="請選擇高麗菜黑豬水餃數量", contents=bubble)
    )


def send_qty_b_flex(reply_token, qty_a):
    remaining = MAX_PACKS - qty_a
    buttons = []
    for i in range(0, remaining + 1):
        buttons.append(
            ButtonComponent(
                action=PostbackAction(label=str(i), data=f"action=qty_b&value={i}"),
                style="primary" if i > 0 else "secondary",
                color="#5B9BD5" if i > 0 else "#AAAAAA",
                height="sm"
            )
        )
    rows = []
    row = []
    for btn in buttons:
        row.append(btn)
        if len(row) == 4:
            rows.append(BoxComponent(layout="horizontal", contents=row, spacing="sm"))
            row = []
    if row:
        rows.append(BoxComponent(layout="horizontal", contents=row, spacing="sm"))

    bubble = BubbleContainer(
        body=BoxComponent(
            layout="vertical",
            contents=[
                TextComponent(text="韭菜黑豬水餃", weight="bold", size="lg", color="#333333"),
                TextComponent(
                    text=f"已選高麗菜 {qty_a} 包，最多還可選 {remaining} 包",
                    size="sm", color="#888888", margin="sm"
                ),
                SeparatorComponent(margin="md"),
                *rows,
                make_cancel_button()
            ],
            spacing="md"
        )
    )
    line_bot_api.reply_message(
        reply_token,
        FlexSendMessage(alt_text="請選擇韭菜黑豬水餃數量", contents=bubble)
    )


def send_identity_check_flex(reply_token, name, phone, address):
    bubble = BubbleContainer(
        body=BoxComponent(
            layout="vertical",
            contents=[
                TextComponent(text="📋 找到上次的資料", weight="bold", size="lg", color="#333333"),
                SeparatorComponent(margin="md"),
                TextComponent(text=f"姓名：{name}", size="sm", color="#555555", margin="md"),
                TextComponent(text=f"電話：{phone}", size="sm", color="#555555", margin="sm"),
                TextComponent(text=f"地址：{address}", size="sm", color="#555555", margin="sm", wrap=True),
                SeparatorComponent(margin="md"),
                TextComponent(text="請問要使用這筆資料嗎？", size="sm", color="#888888", margin="md"),
            ]
        ),
        footer=BoxComponent(
            layout="vertical",
            spacing="sm",
            contents=[
                ButtonComponent(
                    action=PostbackAction(label="✅ 沿用上次資料", data="action=use_saved"),
                    style="primary",
                    color="#FF6B6B"
                ),
                ButtonComponent(
                    action=PostbackAction(label="✏️ 重新填寫", data="action=reenter"),
                    style="secondary"
                ),
                make_cancel_button()
            ]
        )
    )
    line_bot_api.reply_message(
        reply_token,
        FlexSendMessage(alt_text="找到上次資料，請確認", contents=bubble)
    )


def send_delivery_time_flex(reply_token):
    bubble = BubbleContainer(
        body=BoxComponent(
            layout="vertical",
            contents=[
                TextComponent(text="🚚 到貨時間偏好", weight="bold", size="lg", color="#333333"),
                TextComponent(text="請選擇您方便收貨的時間", size="sm", color="#888888", margin="sm"),
                SeparatorComponent(margin="md"),
            ]
        ),
        footer=BoxComponent(
            layout="vertical",
            spacing="sm",
            contents=[
                ButtonComponent(
                    action=PostbackAction(label="平日（週一至週五）", data="action=delivery&value=平日"),
                    style="primary",
                    color="#FF6B6B"
                ),
                ButtonComponent(
                    action=PostbackAction(label="禮拜六", data="action=delivery&value=禮拜六"),
                    style="primary",
                    color="#FF6B6B"
                ),
                ButtonComponent(
                    action=PostbackAction(label="皆可", data="action=delivery&value=皆可"),
                    style="secondary"
                ),
                make_cancel_button()
            ]
        )
    )
    line_bot_api.reply_message(
        reply_token,
        FlexSendMessage(alt_text="請選擇到貨時間", contents=bubble)
    )


def send_order_summary_flex(reply_token, state):
    qty_a = state.get("qty_a", 0)
    qty_b = state.get("qty_b", 0)
    name = state.get("name", "")
    phone = state.get("phone", "")
    address = state.get("address", "")
    delivery_time = state.get("delivery_time", "")
    subtotal, shipping, total = calc_order(qty_a, qty_b)

    content_lines = []
    if qty_a > 0:
        content_lines.append(f"高麗菜黑豬水餃 x{qty_a} 包")
    if qty_b > 0:
        content_lines.append(f"韭菜黑豬水餃 x{qty_b} 包")
    content_str = "\n".join(content_lines)
    shipping_str = "免運 🎉" if shipping == 0 else f"NT${shipping}"

    bubble = BubbleContainer(
        body=BoxComponent(
            layout="vertical",
            contents=[
                TextComponent(text="📦 訂單確認", weight="bold", size="xl", color="#333333"),
                SeparatorComponent(margin="md"),
                TextComponent(text="🛒 商品", weight="bold", size="sm", color="#555555", margin="md"),
                TextComponent(text=content_str, size="sm", color="#333333", margin="sm", wrap=True),
                SeparatorComponent(margin="md"),
                TextComponent(text="👤 收件資訊", weight="bold", size="sm", color="#555555", margin="md"),
                TextComponent(text=f"姓名：{name}", size="sm", color="#333333", margin="sm"),
                TextComponent(text=f"電話：{phone}", size="sm", color="#333333", margin="sm"),
                TextComponent(text=f"地址：{address}", size="sm", color="#333333", margin="sm", wrap=True),
                TextComponent(text=f"到貨時間：{delivery_time}", size="sm", color="#333333", margin="sm"),
                SeparatorComponent(margin="md"),
                TextComponent(text="💰 費用明細", weight="bold", size="sm", color="#555555", margin="md"),
                BoxComponent(
                    layout="horizontal",
                    contents=[
                        TextComponent(text="小計", size="sm", color="#555555"),
                        TextComponent(text=f"NT${subtotal}", size="sm", color="#333333", align="end")
                    ],
                    margin="sm"
                ),
                BoxComponent(
                    layout="horizontal",
                    contents=[
                        TextComponent(text="運費", size="sm", color="#555555"),
                        TextComponent(text=shipping_str, size="sm", color="#333333", align="end")
                    ],
                    margin="sm"
                ),
                BoxComponent(
                    layout="horizontal",
                    contents=[
                        TextComponent(text="總計", size="sm", color="#333333", weight="bold"),
                        TextComponent(text=f"NT${total}", size="sm", color="#FF6B6B", weight="bold", align="end")
                    ],
                    margin="sm"
                ),
                SeparatorComponent(margin="md"),
                TextComponent(text="💳 付款資訊", weight="bold", size="sm", color="#555555", margin="md"),
                TextComponent(
                    text="中國信託銀行（822）\n頭份分行\n帳號：370540364486\n戶名：徐志帆",
                    size="sm",
                    color="#1E90FF",
                    margin="sm",
                    wrap=True
                ),
                TextComponent(
                    text=f"請轉帳 NT${total}，完成後請告知匯款帳號後5碼",
                    size="sm",
                    color="#FF6B6B",
                    margin="sm",
                    wrap=True
                ),
            ]
        ),
        footer=BoxComponent(
            layout="vertical",
            spacing="sm",
            contents=[
                ButtonComponent(
                    action=PostbackAction(label="✅ 確認送出訂單", data="action=confirm_order"),
                    style="primary",
                    color="#FF6B6B"
                ),
                make_cancel_button()
            ]
        )
    )
    line_bot_api.reply_message(
        reply_token,
        FlexSendMessage(alt_text="請確認您的訂單", contents=bubble)
    )


def notify_owner(state, line_display_name="", order_id=None):
    print(f"[notify_owner] ===== 開始執行 =====")
    print(f"[notify_owner] OWNER_ID = {OWNER_ID}")
    print(f"[notify_owner] order_id = {order_id}")
    print(f"[notify_owner] line_display_name = {line_display_name}")

    qty_a = state.get("qty_a", 0)
    qty_b = state.get("qty_b", 0)
    subtotal, shipping, total = calc_order(qty_a, qty_b)

    content_lines = []
    if qty_a > 0:
        content_lines.append(f"高麗菜黑豬水餃 x{qty_a} 包")
    if qty_b > 0:
        content_lines.append(f"韭菜黑豬水餃 x{qty_b} 包")
    content_str = "、".join(content_lines)

    bubble = BubbleContainer(
        body=BoxComponent(
            layout="vertical",
            contents=[
                TextComponent(text="🔔 新訂單通知", weight="bold", size="xl", color="#333333"),
                SeparatorComponent(margin="md"),
                TextComponent(text=f"💬 LINE：{line_display_name}", size="sm", color="#555555", margin="md"),
                TextComponent(text=f"👤 姓名：{state.get('name')}", size="sm", color="#555555", margin="sm"),
                TextComponent(text=f"📞 電話：{state.get('phone')}", size="sm", color="#555555", margin="sm"),
                TextComponent(
                    text=f"📍 地址：{state.get('address')}",
                    size="sm", color="#555555", margin="sm", wrap=True
                ),
                TextComponent(
                    text=f"🚚 到貨：{state.get('delivery_time')}",
                    size="sm", color="#555555", margin="sm"
                ),
                SeparatorComponent(margin="md"),
                TextComponent(
                    text=f"🛒 {content_str}",
                    size="sm", color="#333333", margin="md", wrap=True
                ),
                BoxComponent(
                    layout="horizontal", margin="sm",
                    contents=[
                        TextComponent(text="小計", size="sm", color="#888888"),
                        TextComponent(text=f"NT${subtotal}", size="sm", color="#333333", align="end"),
                    ]
                ),
                BoxComponent(
                    layout="horizontal", margin="sm",
                    contents=[
                        TextComponent(text="運費", size="sm", color="#888888"),
                        TextComponent(
                            text="免運 🎉" if shipping == 0 else f"NT${shipping}",
                            size="sm", color="#333333", align="end"
                        ),
                    ]
                ),
                BoxComponent(
                    layout="horizontal", margin="sm",
                    contents=[
                        TextComponent(text="總計", size="sm", color="#333333", weight="bold"),
                        TextComponent(
                            text=f"NT${total}",
                            size="sm", color="#FF6B6B", weight="bold", align="end"
                        ),
                    ]
                ),
                SeparatorComponent(margin="md"),
                TextComponent(
                    text=f"💰 匯款後5碼：{state.get('last_5_digits', '待確認')}",
                    size="sm", color="#FF6B6B", margin="md", weight="bold"
                ),
            ]
        ),
        footer=BoxComponent(
            layout="vertical",
            contents=[
                ButtonComponent(
                    action=PostbackAction(
                        label="✅ 確認收款",
                        data=f"action=confirm_payment&order_id={order_id}"
                    ),
                    style="primary",
                    color="#28A745"
                )
            ]
        )
    )

    try:
        print(f"[notify_owner] 準備發送 push_message 給 {OWNER_ID}...")
        line_bot_api.push_message(
            OWNER_ID,
            FlexSendMessage(alt_text="🔔 新訂單通知", contents=bubble)
        )
        print(f"[notify_owner] ✅ 發送成功！")
    except Exception as e:
        print(f"[notify_owner] ❌ 發送失敗：{e}")
        traceback.print_exc()


def save_order(user_id, state, last_5_digits=""):
    qty_a = state.get("qty_a", 0)
    qty_b = state.get("qty_b", 0)
    subtotal, shipping, total = calc_order(qty_a, qty_b)

    result = supabase.table("orders").insert({
        "user_id": user_id,
        "name": state.get("name"),
        "phone": state.get("phone"),
        "address": state.get("address"),
        "delivery_time": state.get("delivery_time"),
        "qty_a": qty_a,
        "qty_b": qty_b,
        "subtotal": subtotal,
        "shipping": shipping,
        "total": total,
        "exported": False,
        "status": "pending",
        "last_5_digits": last_5_digits
    }).execute()

    supabase.table("user_profiles").upsert({
        "user_id": user_id,
        "name": state.get("name"),
        "phone": state.get("phone"),
        "address": state.get("address"),
        "updated_at": datetime.datetime.utcnow().isoformat()
    }).execute()

    return result.data[0]["id"]


def export_orders_to_excel():
    try:
        result = supabase.table("orders") \
            .select("*") \
            .eq("exported", False) \
            .eq("status", "paid") \
            .order("created_at", desc=False) \
            .execute()

        orders = result.data

        if not orders:
            return None, "目前沒有已收款的訂單可匯出 📭"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "訂單列表"

        headers = ["溫層", "收件人姓名", "收件人電話", "收件人手機", "收件人地址", "寄件日期", "包裹內容", "品名類別", "希望送達時段"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        ws.row_dimensions[1].height = 20

        for row_idx, order in enumerate(orders, start=2):
            content_parts = []
            if order.get("qty_a", 0) > 0:
                content_parts.append(f"高麗菜黑豬水餃 x{order['qty_a']}包")
            if order.get("qty_b", 0) > 0:
                content_parts.append(f"韭菜黑豬水餃 x{order['qty_b']}包")
            content_str = "、".join(content_parts)

            delivery_map = {
                "平日": "週一至週五",
                "禮拜六": "週六",
                "皆可": "不限"
            }
            delivery_display = delivery_map.get(
                order.get("delivery_time", ""),
                order.get("delivery_time", "")
            )

            row_data = [
                "冷凍",
                order.get("name", ""),
                "",
                order.get("phone", ""),
                order.get("address", ""),
                "",
                content_str,
                "一般食品",
                delivery_display,
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(
                        start_color="DCE6F1",
                        end_color="DCE6F1",
                        fill_type="solid"
                    )

        col_widths = [10, 12, 12, 15, 40, 14, 35, 12, 14]
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        ws.freeze_panes = "A2"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"orders_{timestamp}.xlsx"

        supabase.storage.from_("exports").upload(
            path=file_name,
            file=buffer.getvalue(),
            file_options={
                "content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
        )

        public_url = supabase.storage.from_("exports").get_public_url(file_name)

        order_ids = [order["id"] for order in orders]
        supabase.table("orders") \
            .update({"exported": True}) \
            .in_("id", order_ids) \
            .execute()

        return public_url, f"✅ 成功匯出 {len(orders)} 筆訂單！"

    except Exception as e:
        print(f"[export] 匯出錯誤: {e}")
        traceback.print_exc()
        return None, f"❌ 匯出失敗：{str(e)}"


@app.route("/callback", methods=["POST"])
def callback():
    signature = request.headers["X-Line-Signature"]
    body = request.get_data(as_text=True)
    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        abort(400)
    return "OK"


@handler.add(MessageEvent, message=TextMessage)
def handle_message(event):
    user_id = event.source.user_id
    user_message = event.message.text.strip()

    print(f"[handle_message] user_id={user_id}, message={user_message}")

    source_type = event.source.type
    if source_type == "group":
        source_id = event.source.group_id
    elif source_type == "room":
        source_id = event.source.room_id
    else:
        source_id = user_id

    # 取消關鍵字（僅限私訊）
    if source_type != "group":
        cancel_keywords = ["取消", "離開", "掰掰"]
        if user_message in cancel_keywords:
            reset_state(user_id)
            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(text="已取消訂單，歡迎再次訂購！🥟")
            )
            return

    # 查詢自己的 LINE ID
    if user_message == "我的ID":
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text=f"你的 User ID 是：\n{user_id}")
        )
        return

    # 測試通知（僅限 OWNER）
    if user_message == "測試通知" and user_id == OWNER_ID:
        print(f"[測試通知] 由 {user_id} 觸發")
        test_state = {
            "qty_a": 2,
            "qty_b": 1,
            "name": "測試用戶",
            "phone": "0912345678",
            "address": "台灣某地某路某號",
            "delivery_time": "平日",
            "last_5_digits": "99999"
        }
        notify_owner(test_state, line_display_name="測試帳號", order_id="TEST-001")
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text="✅ 測試通知已發送，請確認是否收到！")
        )
        return

    # 匯出 Excel（僅限 OWNER）
    if user_message == "匯出" and (user_id == OWNER_ID or source_id == OWNER_ID):
        url, msg = export_orders_to_excel()
        reply_text = f"{msg}\n\n📥 下載連結：\n{url}" if url else msg
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text=reply_text)
        )
        return

    # 群組訊息不處理後續流程
    if source_type == "group":
        return

    state = user_states.get(user_id, {})
    step = state.get("step", 0)

    print(f"[handle_message] step={step}")

    # Step 0：歡迎畫面
    if step == 0:
        send_welcome(event.reply_token)
        return

    # Step 4：輸入姓名
    if step == 4:
        state["name"] = user_message
        state["step"] = 5
        user_states[user_id] = state
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text="請輸入您的手機號碼（格式：09xxxxxxxx）")
        )
        return

    # Step 5：輸入電話
    if step == 5:
        if not re.match(r'^09\d{8}$', user_message):
            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(text="❗ 手機號碼格式錯誤，請重新輸入（格式：09xxxxxxxx）")
            )
            return
        state["phone"] = user_message
        state["step"] = 6
        user_states[user_id] = state
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text="請輸入收件地址")
        )
        return

    # Step 6：輸入地址
    if step == 6:
        state["address"] = user_message
        state["step"] = 7
        user_states[user_id] = state
        send_delivery_time_flex(event.reply_token)
        return

    # Step 9：輸入匯款後5碼
    if step == 9:
        print(f"[step9] 收到後5碼輸入：{user_message}")
        print(f"[step9] state = {state}")

        if not re.match(r'^\d{5}$', user_message):
            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(text="❗ 請輸入 5 位數字，例如：12345")
            )
            return

        order_id = state.get("order_id")
        line_display_name = state.get("line_display_name", "（無法取得）")

        print(f"[step9] order_id={order_id}, line_display_name={line_display_name}")

        # 更新資料庫的後5碼
        supabase.table("orders") \
            .update({"last_5_digits": user_message}) \
            .eq("id", order_id) \
            .execute()

        state["last_5_digits"] = user_message

        # 通知店家
        notify_owner(state, line_display_name, order_id)

        reset_state(user_id)

        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(
                text=(
                    f"✅ 收到您的後5碼：{user_message}\n\n"
                    f"確認收款後我們會盡快安排出貨，謝謝您 🥟"
                )
            )
        )
        return

    # 其他情況
    line_bot_api.reply_message(
        event.reply_token,
        TextSendMessage(text="請使用按鈕操作，或輸入「取消」結束訂購 🥟")
    )


@handler.add(PostbackEvent)
def handle_postback(event):
    user_id = event.source.user_id
    data = event.postback.data
    params = dict(item.split("=") for item in data.split("&"))
    action = params.get("action")

    print(f"[handle_postback] user_id={user_id}, action={action}, params={params}")

    # 取消
    if action == "cancel":
        reset_state(user_id)
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text="已取消訂單，歡迎再次訂購！🥟")
        )
        return

    # 開始訂購
    if action == "start_order":
        user_states[user_id] = {"step": 1}
        send_qty_a_flex(event.reply_token)
        return

    # 確認收款（店家按鈕）
    if action == "confirm_payment":
        order_id = params.get("order_id")
        print(f"[confirm_payment] order_id={order_id}")
        try:
            result = supabase.table("orders") \
                .update({"status": "paid"}) \
                .eq("id", order_id) \
                .execute()

            order = result.data[0]
            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(
                    text=(
                        f"✅ 已確認收款！\n"
                        f"訂單 #{order_id}\n"
                        f"客戶：{order['name']}\n"
                        f"總計：NT${order['total']}\n\n"
                        f"此訂單將在下次匯出時納入 📋"
                    )
                )
            )
        except Exception as e:
            print(f"[confirm_payment] 失敗：{e}")
            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(text=f"❌ 更新失敗：{str(e)}")
            )
        return

    state = user_states.get(user_id, {})
    step = state.get("step", 0)

    print(f"[handle_postback] step={step}")

    # Step 1：選擇高麗菜數量
    if action == "qty_a" and step == 1:
        qty_a = int(params.get("value", 0))
        state["qty_a"] = qty_a
        state["step"] = 2
        user_states[user_id] = state
        send_qty_b_flex(event.reply_token, qty_a)
        return

    # Step 2：選擇韭菜數量
    if action == "qty_b" and step == 2:
        qty_a = state.get("qty_a", 0)
        qty_b = int(params.get("value", 0))
        total_packs = qty_a + qty_b

        if total_packs < MIN_PACKS:
            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(
                    text=f"❗ 最少需訂購 {MIN_PACKS} 包，目前共 {total_packs} 包，請重新選擇。"
                )
            )
            line_bot_api.push_message(
                user_id,
                FlexSendMessage(
                    alt_text="請選擇韭菜黑豬水餃數量",
                    contents=_build_qty_b_bubble(qty_a)
                )
            )
            return

        state["qty_b"] = qty_b
        state["step"] = 3
        user_states[user_id] = state

        profile_result = supabase.table("user_profiles") \
            .select("*") \
            .eq("user_id", user_id) \
            .execute()

        if profile_result.data:
            profile = profile_result.data[0]
            send_identity_check_flex(
                event.reply_token,
                profile["name"],
                profile["phone"],
                profile["address"]
            )
        else:
            state["step"] = 4
            user_states[user_id] = state
            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(text="請輸入您的姓名")
            )
        return

    # Step 3：沿用舊資料
    if action == "use_saved" and step == 3:
        profile_result = supabase.table("user_profiles") \
            .select("*") \
            .eq("user_id", user_id) \
            .execute()

        if profile_result.data:
            profile = profile_result.data[0]
            state["name"] = profile["name"]
            state["phone"] = profile["phone"]
            state["address"] = profile["address"]
            state["step"] = 7
            user_states[user_id] = state
            send_delivery_time_flex(event.reply_token)
        return

    # Step 3：重新填寫
    if action == "reenter" and step == 3:
        state["step"] = 4
        user_states[user_id] = state
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text="請輸入您的姓名")
        )
        return

    # Step 7：選擇到貨時間
    if action == "delivery" and step == 7:
        state["delivery_time"] = params.get("value", "皆可")
        state["step"] = 8
        user_states[user_id] = state
        send_order_summary_flex(event.reply_token, state)
        return

    # Step 8：確認送出訂單
    if action == "confirm_order" and step == 8:
        print(f"[confirm_order] 開始儲存訂單")
        try:
            try:
                profile = line_bot_api.get_profile(user_id)
                line_display_name = profile.display_name
                print(f"[confirm_order] 取得顯示名稱：{line_display_name}")
            except Exception as e:
                print(f"[confirm_order] 無法取得顯示名稱：{e}")
                line_display_name = "（無法取得）"

            order_id = save_order(user_id, state, last_5_digits="")
            print(f"[confirm_order] 訂單儲存成功，order_id={order_id}")

            state["step"] = 9
            state["order_id"] = order_id
            state["line_display_name"] = line_display_name
            user_states[user_id] = state

            subtotal, shipping, total = calc_order(
                state.get("qty_a", 0), state.get("qty_b", 0)
            )

            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(
                    text=(
                        f"✅ 訂單已送出！感謝您的訂購 🥟\n\n"
                        f"💳 請轉帳 NT${total} 至：\n"
                        f"中國信託（822）頭份分行\n"
                        f"帳號：370540364486\n"
                        f"戶名：徐志帆\n\n"
                        f"轉帳完成後，請回覆匯款帳號後 5 碼 🔢"
                    )
                )
            )
        except Exception as e:
            print(f"[confirm_order] 訂單儲存失敗: {e}")
            traceback.print_exc()
            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(text=f"❌ 訂單送出失敗，請稍後再試。\n錯誤：{str(e)}")
            )
        return


def _build_qty_b_bubble(qty_a):
    remaining = MAX_PACKS - qty_a
    buttons = []
    for i in range(0, remaining + 1):
        buttons.append(
            ButtonComponent(
                action=PostbackAction(label=str(i), data=f"action=qty_b&value={i}"),
                style="primary" if i > 0 else "secondary",
                color="#5B9BD5" if i > 0 else "#AAAAAA",
                height="sm"
            )
        )
    rows = []
    row = []
    for btn in buttons:
        row.append(btn)
        if len(row) == 4:
            rows.append(BoxComponent(layout="horizontal", contents=row, spacing="sm"))
            row = []
    if row:
        rows.append(BoxComponent(layout="horizontal", contents=row, spacing="sm"))

    return BubbleContainer(
        body=BoxComponent(
            layout="vertical",
            contents=[
                TextComponent(text="韭菜黑豬水餃", weight="bold", size="lg", color="#333333"),
                TextComponent(
                    text=f"已選高麗菜 {qty_a} 包，最多還可選 {remaining} 包",
                    size="sm", color="#888888", margin="sm"
                ),
                SeparatorComponent(margin="md"),
                *rows,
                make_cancel_button()
            ],
            spacing="md"
        )
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
